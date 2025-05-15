from datetime import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

def generate_monthly_report_pdf(data, output_path):
    """Generate a PDF report for the user's investment portfolio"""
    try:
        # Create PDF object
        pdf = FPDF()
        pdf.add_page()
        
        # Set fonts
        pdf.set_font('Arial', 'B', 16)
        
        # Title
        pdf.cell(0, 10, f"Investment Portfolio Report - {data['month']} {data['year']}", 0, 1, 'C')
        pdf.cell(0, 10, f"Generated for: {data['username']}", 0, 1, 'C')
        pdf.ln(10)
        
        # Summary
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, "Summary", 0, 1)
        
        pdf.set_font('Arial', '', 12)
        pdf.cell(100, 10, "Total Investment:", 0, 0)
        pdf.cell(0, 10, f"₹ {data['total_investment']:.2f}", 0, 1)
        
        pdf.cell(100, 10, "Current Portfolio Value:", 0, 0)
        pdf.cell(0, 10, f"₹ {data['total_current_value']:.2f}", 0, 1)
        
        pdf.cell(100, 10, "Net Gain/Loss:", 0, 0)
        pdf.cell(0, 10, f"₹ {data['net_gain_loss']:.2f} ({data['net_gain_loss_percent']:.2f}%)", 0, 1)
        
        pdf.cell(100, 10, "Top Performing Stock:", 0, 0)
        pdf.cell(0, 10, f"{data['top_gainer']['symbol']} ({data['top_gainer']['gain_percent']:.2f}%)", 0, 1)
        
        pdf.cell(100, 10, "Worst Performing Stock:", 0, 0)
        pdf.cell(0, 10, f"{data['top_loser']['symbol']} ({data['top_loser']['loss_percent']:.2f}%)", 0, 1)
        
        pdf.ln(10)
        
        # Portfolio details
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, "Portfolio Details", 0, 1)
        
        # Table header
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(25, 10, "Symbol", 1, 0, 'C')
        pdf.cell(25, 10, "Quantity", 1, 0, 'C')
        pdf.cell(30, 10, "Buy Price", 1, 0, 'C')
        pdf.cell(30, 10, "Current Price", 1, 0, 'C')
        pdf.cell(30, 10, "Investment", 1, 0, 'C')
        pdf.cell(30, 10, "Current Value", 1, 0, 'C')
        pdf.cell(25, 10, "Gain/Loss %", 1, 1, 'C')
        
        # Table rows
        pdf.set_font('Arial', '', 10)
        for item in data['portfolio']:
            pdf.cell(25, 10, item['symbol'], 1, 0)
            pdf.cell(25, 10, f"{item['quantity']:.2f}", 1, 0, 'R')
            pdf.cell(30, 10, f"₹ {item['buy_price']:.2f}", 1, 0, 'R')
            pdf.cell(30, 10, f"₹ {item['current_price']:.2f}", 1, 0, 'R')
            pdf.cell(30, 10, f"₹ {item['investment']:.2f}", 1, 0, 'R')
            pdf.cell(30, 10, f"₹ {item['current_value']:.2f}", 1, 0, 'R')
            pdf.cell(25, 10, f"{item['gain_loss_percent']:.2f}%", 1, 1, 'R')
        
        # Footer
        pdf.ln(10)
        pdf.set_font('Arial', 'I', 8)
        pdf.cell(0, 10, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 0, 'C')
        
        # Output to file
        pdf.output(output_path)
        
        return True
    except Exception as e:
        logging.error(f"Error generating PDF report: {str(e)}")
        return False

def generate_monthly_report_excel(data, output_path):
    """Generate an Excel report for the user's investment portfolio"""
    try:
        # Create workbook and select active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Investment Report {data['month']}"
        
        # Define styles
        title_font = Font(name='Calibri', size=16, bold=True)
        header_font = Font(name='Calibri', size=12, bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"Investment Portfolio Report - {data['month']} {data['year']}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A2'] = f"Generated for: {data['username']}"
        ws['A2'].font = header_font
        ws.merge_cells('A2:G2')
        
        # Summary
        ws['A4'] = "Summary"
        ws['A4'].font = header_font
        ws.merge_cells('A4:G4')
        
        ws['A5'] = "Total Investment:"
        ws['B5'] = data['total_investment']
        ws['B5'].number_format = '₹#,##0.00'
        
        ws['A6'] = "Current Portfolio Value:"
        ws['B6'] = data['total_current_value']
        ws['B6'].number_format = '₹#,##0.00'
        
        ws['A7'] = "Net Gain/Loss:"
        ws['B7'] = data['net_gain_loss']
        ws['B7'].number_format = '₹#,##0.00'
        ws['C7'] = f"({data['net_gain_loss_percent']:.2f}%)"
        
        ws['A8'] = "Top Performing Stock:"
        ws['B8'] = f"{data['top_gainer']['symbol']} ({data['top_gainer']['gain_percent']:.2f}%)"
        
        ws['A9'] = "Worst Performing Stock:"
        ws['B9'] = f"{data['top_loser']['symbol']} ({data['top_loser']['loss_percent']:.2f}%)"
        
        # Portfolio details - Table headers
        ws['A11'] = "Portfolio Details"
        ws['A11'].font = header_font
        ws.merge_cells('A11:G11')
        
        headers = ["Symbol", "Quantity", "Buy Price", "Current Price", "Investment", "Current Value", "Gain/Loss %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Portfolio details - Data
        for row, item in enumerate(data['portfolio'], 13):
            ws.cell(row=row, column=1).value = item['symbol']
            ws.cell(row=row, column=2).value = item['quantity']
            ws.cell(row=row, column=3).value = item['buy_price']
            ws.cell(row=row, column=3).number_format = '₹#,##0.00'
            ws.cell(row=row, column=4).value = item['current_price']
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=5).value = item['investment']
            ws.cell(row=row, column=5).number_format = '₹#,##0.00'
            ws.cell(row=row, column=6).value = item['current_value']
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
            ws.cell(row=row, column=7).value = f"{item['gain_loss_percent']:.2f}%"
            
            # Apply borders to cells
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Footer
        footer_row = 13 + len(data['portfolio']) + 2
        ws.cell(row=footer_row, column=1).value = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells(f'A{footer_row}:G{footer_row}')
        ws.cell(row=footer_row, column=1).font = Font(name='Calibri', size=8, italic=True)
        ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='center')
        
        # Save workbook
        wb.save(output_path)
        
        return True
    except Exception as e:
        logging.error(f"Error generating Excel report: {str(e)}")
        return False
